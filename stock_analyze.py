import streamlit as st
import yfinance as yf
import pandas as pd
import pandas_ta as ta
import time
import random
import requests
import urllib3
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# --- 基礎配置 ---
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
st.set_page_config(page_title="台股多因子決策系統 (加碼止損版)", layout="wide")

# --- 1. 全面獲取股票代碼 (全面模式) ---
@st.cache_data(ttl=86400)
def get_full_market_tickers():
    url = "https://isin.twse.com.tw/isin/C_public.jsp?strMode=2"
    try:
        res = requests.get(url, timeout=10, verify=False, headers={'User-Agent': 'Mozilla/5.0'})
        res.encoding = 'big5'
        df = pd.read_html(res.text)[0]
        df.columns = df.iloc[0]
        df = df[df['有價證券代號及名稱'].str.contains("　", na=False)] # 注意：這裡是全形空格
        tickers = [f"{t.split('　')[0].strip()}.TW" for t in df['有價證券代號及名稱'] if len(t.split('　')[0].strip()) == 4]
        if len(tickers) > 800: return tickers
    except: pass
    # 如果爬蟲失敗，回傳保底清單
    return [f"{i:04d}.TW" for i in range(1101, 9999)]

# --- 2. 輸出到 Excel ---
def export_to_excel(data, filename):
    """將掃描結果輸出到 Excel 檔案"""
    wb = Workbook()
    sheet = wb.active
    sheet.title = "市場掃描結果"
    
    # 設定標題 (保持與你原本設計一致)
    headers = ["日期", "股票代碼", "實際收盤價"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 寫入資料
    current_date = datetime.now().strftime("%Y-%m-%d")
    for row_num, row_data in enumerate(data, 2):
        sheet.cell(row=row_num, column=1, value=current_date)
        sheet.cell(row=row_num, column=2, value=row_data['股票代號'].replace('.TW', ''))
        sheet.cell(row=row_num, column=3, value=row_data['收盤價'])
    
    # 調整欄寬
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 15
    
    wb.save(filename)
    return filename

# --- 3. 交易決策邏輯 (整合回測標準) ---
def analyze_stock_advanced(ticker, weights, params):
    try:
        df = yf.download(ticker, period="60d", interval="1d", progress=False, auto_adjust=True)
        if df.empty or len(df) < 20: return None
        if isinstance(df.columns, pd.MultiIndex): df.columns = df.columns.get_level_values(0)

        df['RSI'] = ta.rsi(df['Close'], length=14)
        df['MA5'] = ta.sma(df['Close'], length=5)
        df['MA10'] = ta.sma(df['Close'], length=10)

        curr, prev = df.iloc[-1], df.iloc[-2]
        c_price = float(curr['Close'])
        c_rsi = float(curr['RSI'])
        
        # 評分邏輯
        score = 0
        if c_rsi < 30: score += weights['rsi']
        if float(prev['MA5']) < float(prev['MA10']) and float(curr['MA5']) > float(curr['MA10']): score += weights['ma']
        chg = ((c_price - float(prev['Close'])) / float(prev['Close'])) * 100
        if abs(chg) >= 7.0: score += weights['vol']
        if float(curr['Volume']) > df['Volume'].mean() * 2: score += weights['vxx']

        # 動作判定
        action = "觀望"
        
        # 買入判定
        if score >= params['buy_threshold']:
            action = "🟢 建議買入"
        # RSI 過熱
        elif c_rsi > params['overbought_rsi']:
            action = "🔵 建議賣出"
        # 部分調節
        elif c_rsi > params['profit_take_rsi']:
            action = "🟠 部分調節"

        return {
            "代碼": ticker, "總分": score, "現價": round(c_price, 2),
            "RSI": round(c_rsi, 1), "建議動作": action
        }
    except: return None

# --- UI 導航 ---
page = st.sidebar.radio("功能選單", ["1. 全市場資金選股", "2. 進階決策中心"])

# 參數設定
st.sidebar.divider()
st.sidebar.header("⚙️ 交易策略參數")
pt_rsi = st.sidebar.slider("部分調節 RSI", 40, 70, 60)
ob_rsi = st.sidebar.slider("獲利清倉 RSI", 70, 95, 80)

# --- 頁面 1：選股 ---
if page == "1. 全市場資金選股":
    st.title("🏆 全市場資金指標排行")
    if st.button("🚀 執行深度掃描"):
        all_list = get_full_market_tickers()
        res_rank = []
        p_bar = st.progress(0, text="分批下載中...")
        
        batch_size = 50
        for i in range(0, len(all_list), batch_size):
            batch = all_list[i : i + batch_size]
            try:
                data = yf.download(batch, period="2d", group_by='ticker', threads=True, progress=False)
                for t in batch:
                    try:
                        t_df = data[t].dropna() if isinstance(data.columns, pd.MultiIndex) else data.dropna()
                        if not t_df.empty:
                            last = t_df.iloc[-1]
                            val = (float(last['Close']) * float(last['Volume'])) / 1e8
                            res_rank.append({"股票代號": t, "收盤價": float(last['Close']), "成交值(億)": val})
                    except: continue
            except: pass
            p_bar.progress(min((i + batch_size) / len(all_list), 1.0))
            time.sleep(random.uniform(0.5, 1.0))
        
        if res_rank:
            top_100 = pd.DataFrame(res_rank).sort_values("成交值(億)", ascending=False).head(100)
            st.session_state.top_100_list = top_100['股票代號'].tolist()
            st.dataframe(top_100, use_container_width=True)
            
            # --- 修正後的 Excel 輸出路徑 ---
            # 移除了 /mnt/user-data/outputs/ 這種絕對路徑，改用當前目錄
            excel_filename = f"market_scan_{datetime.now().strftime('%Y%m%d')}.xlsx"
            export_to_excel(res_rank, excel_filename)
            
            st.success(f"✅ 已將掃描結果輸出至 Excel 檔案")
            
            # 提供下載連結
            with open(excel_filename, "rb") as f:
                st.download_button(
                    label="📥 下載 Excel 檔案",
                    data=f,
                    file_name=excel_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

# --- 頁面 2：決策 ---
elif page == "2. 進階決策中心":
    st.title("🛡️ 進階量化決策中心")
    if 'top_100_list' not in st.session_state:
        st.warning("請先執行第一頁掃描。")
    else:
        weights = {'rsi': 40, 'ma': 30, 'vol': 20, 'vxx': 10}
        params = {
            'profit_take_rsi': pt_rsi, 'overbought_rsi': ob_rsi, 'buy_threshold': 30
        }
        
        signals = []
        p_check = st.progress(0, text="計算指標中...")
        for idx, t in enumerate(st.session_state.top_100_list):
            res = analyze_stock_advanced(t, weights, params)
            if res: signals.append(res)
            p_check.progress((idx + 1) / 100)
        
        if signals:
            st.dataframe(pd.DataFrame(signals).sort_values("總分", ascending=False), use_container_width=True)
